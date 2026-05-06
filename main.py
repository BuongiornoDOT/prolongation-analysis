import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

#путь к исходным данным
prolongations = pd.read_csv("data/prolongations.csv")
financial = pd.read_csv("data/financial_data.csv")

all_months = ["Ноябрь 2022",
         "Декабрь 2022",
         "Январь 2023",
         "Февраль 2023",
         "Март 2023",
         "Апрель 2023",
         "Май 2023",
         "Июнь 2023",
         "Июль 2023",
         "Август 2023",
         "Сентябрь 2023",
         "Октябрь 2023",
         "Ноябрь 2023",
         "Декабрь 2023"
        ]
report_months = ["Январь 2023",
                  "Февраль 2023",
                  "Март 2023",
                  "Апрель 2023",
                  "Май 2023",
                  "Июнь 2023",
                  "Июль 2023",
                  "Август 2023",
                  "Сентябрь 2023",
                  "Октябрь 2023",
                  "Ноябрь 2023",
                  "Декабрь 2023"
                  ]
results = []

#подготовка рабочей таблицы проектов
am_data = prolongations[["id", "month", "AM"]].drop_duplicates(subset="id", keep="last")
financial_list = financial.groupby("id", as_index=False)[all_months].sum()
project_list = financial_list.merge(am_data, on="id", how="left")
project_list = project_list[project_list["month"].notna() & project_list["AM"].notna()]
project_list["month"] = project_list["month"].str.title()
am_list = project_list["AM"].dropna().unique()

#очистка таблицы
for month in all_months:
    project_list[month] = project_list[month].str.replace(" ", "")
    project_list[month] = project_list[month].str.replace("\xa0", "")
    project_list[month] = project_list[month].str.replace(",", ".")
    project_list[month] = pd.to_numeric(project_list[month], errors="coerce")
    project_list[month] = project_list[month].fillna(0)

# расчет коэффициентов
for month in report_months:
    index = all_months.index(month)
    last_month = all_months[index - 1]
    last_two_month = all_months[index - 2]
    for am in am_list:
        #расчет пролонгации в первый месяц
        am_projects = project_list[project_list["AM"] == am]
        project_last_month = am_projects[am_projects["month"] == last_month]
        base_last_month = project_last_month[last_month].sum()
        prolongation_last_month = project_last_month[month].sum()
        if base_last_month != 0:
            coefficient_last_month = prolongation_last_month / base_last_month
        else:
            coefficient_last_month = None
        #расчет пролонгации во второй месяц
        project_last_two_month = am_projects[am_projects["month"] == last_two_month]
        project_last_two_month = project_last_two_month[project_last_two_month[project_last_month] == 0]
        base_last_two_month = project_last_two_month[last_two_month].sum()
        prolongation_last_two_month = project_last_two_month[month].sum()
        if base_last_two_month != 0:
            coefficient_last_two_month = prolongation_last_two_month / base_last_two_month
        else:
            coefficient_last_two_month = None
        #словарь с результатами цикла
        row = {
                "AM": am,
                "month": month,
                "base_last_month": base_last_month,
                "prolongation_last_month": prolongation_last_month,
                "coefficient_last_month": coefficient_last_month,
                "base_last_two_month": base_last_two_month,
                "prolongation_last_two_month": prolongation_last_two_month,
                "coefficient_last_two_month": coefficient_last_two_month
                }
        results.append(row)

#подготовка рабочей таблицы проектов
monthly_report = pd.DataFrame(results)

wb = Workbook()
wb.remove(wb.active)
header_fill = PatternFill(fill_type="solid", fgColor="FCE5B5")
thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for am in monthly_report["AM"].dropna().unique():
    am_data = monthly_report[monthly_report["AM"] == am].copy()
    safe_title = am
    safe_title = safe_title.replace("/", "-")
    safe_title = safe_title.replace("\\", "-")
    safe_title = safe_title.replace("*", "")
    safe_title = safe_title.replace("?", "")
    safe_title = safe_title.replace(":", "")
    safe_title = safe_title.replace("[", "")
    safe_title = safe_title.replace("]", "")
    safe_title = safe_title[:31]
    ws = wb.create_sheet(title=safe_title)
    ws["A1"] = am
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Месяц"
    ws["B3"] = "Пролонгации в первый месяц"
    ws["E3"] = "Пролонгации через месяц"
    ws.merge_cells("A3:A4")
    ws.merge_cells("B3:D3")
    ws.merge_cells("E3:G3")
    ws["B4"] = "к пролонгации"
    ws["C4"] = "пролонгировано"
    ws["D4"] = "Коэффициент"
    ws["E4"] = "к пролонгации"
    ws["F4"] = "пролонгировано"
    ws["G4"] = "Коэффициент"
    for row in ws["A3:G4"]:
        for cell in row:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    start_row = 5
    for i, (_, row) in enumerate(am_data.iterrows(), start=start_row):
        ws[f"A{i}"] = row["month"]
        ws[f"B{i}"] = row["base_last_month"]
        ws[f"C{i}"] = row["prolongation_last_month"]
        ws[f"D{i}"] = row["coefficient_last_month"]
        ws[f"E{i}"] = row["base_last_two_month"]
        ws[f"F{i}"] = row["prolongation_last_two_month"]
        ws[f"G{i}"] = row["coefficient_last_two_month"]
        for cell in ws[i]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_num in range(start_row, start_row + len(am_data)):
        for col in ["B", "C", "E", "F"]:
            ws[f"{col}{row_num}"].number_format = '# ##0.00'
            for col in ["D", "G"]:
                ws[f"{col}{row_num}"].number_format = '0.00%'
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 36
wb.save("report.xlsx")
